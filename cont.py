import matplotlib.pyplot as plt
from tkinter import *
import openpyxl


with open('out.txt') as f:
    content = f.readlines()

gui = Tk()
gui.geometry('500x110+520+200')

content = [x.strip() for x in content] 
contents=[i.split(',', 1)[0] for i in content]
contentl=[i.split(',', 1)[-1] for i in content]

def ploty():
    plt.plot(contentl,contents)
    plt.ylabel('Bending Stress')
    plt.xlabel('Length (cm)')
    plt.show()

gui.title(': Stress vs Length :')
S = 0
pos = 0
with open('som.txt') as t:
    text = t.readline()
text1 = text.split(" ")

maxmom = 'The maximum bending moment is '+text1[0]+' Nm at position '+text1[4]
maxstr = 'The maximum bending stress is '+text1[3]+' N/m^2 at position '+text1[4]

wb = openpyxl.load_workbook('SOM.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(2,16):
    R = (sheet.cell(row=i, column=2).value)*float(text1[1])/float((text1[0]))
    S = (sheet.cell(row=i, column=2).value)*(float(text1[2])/2)/R
    if(S>float(text1[3])):
        pos = i
        break
if pos>0:
    steel = 'The material whose value of E is greater than '+sheet.cell(row=pos,column=1).value+' can be used.'
else:
    steel = 'No material from the data given can be used'

Label1 = Label(text = maxmom).pack()
Label2 = Label(text = maxstr).pack()
Label3 = Label(text = steel).pack()


mybutton = Button(text='click to Plot...',fg='red',bg='yellow',command = ploty).pack()
gui.mainloop()





